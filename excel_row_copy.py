import tkinter as tk
from tkinter import filedialog
import openpyxl
import os

# Lista para almacenar los datos a copiar
datas = []
 
class FileDialog():
    def __init__(self):
        # Configurar los tipos de archivo permitidos (solo .xlsx)
        tipos_archivo = [("Archivos de Excel", "*.xlsx")]

        # Abrir un cuadro de diálogo para seleccionar un archivo
        self.archivo = filedialog.askopenfilename(filetypes=tipos_archivo)

        # Mostrar el archivo seleccionado
        if self.archivo:
            print(f"Archivo seleccionado: {self.archivo}")
        else:
            print(f"No se seleccionó ningún archivo")


class CopyFile():
    def __init__(self, **kwargs):
        # Crear un nuevo libro de trabajo
        self.wb = openpyxl.Workbook()

        # Crear el nombre del archivo de salida
        filename = kwargs["filename"] + ".xlsx"

        # Abrir el archivo Excel existente para copiar datos
        file_to_copy = openpyxl.load_workbook(kwargs["archivo"])
        dataframe = file_to_copy.active

        # Crear una nueva hoja en el libro de trabajo de salida
        new_file = self.wb.active
        new_file.title = kwargs["title"]

        # Copiar los datos de la hoja de cálculo original a la nueva hoja
        for row in range(1, dataframe.max_row):
            data_for_process = [row, ]
            for data in dataframe.iter_cols(1, dataframe.max_column):
                data_for_process.append(data[row].value)
            datas.append(data_for_process)

        filas_names = []
        rows = [int(x) for x in kwargs["row"].split(",")]
        for i in range(len(rows)):
            nombre_fila = input(f"Escribe el nombre de la fila #{i + 1}: ")
            filas_names.append(nombre_fila)
        print("Creando archivo 75%..................")

        for index, value in enumerate(rows):
            z = 2
            if value == 1:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    z += 1
            if value == 2:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    z += 1
            if value == 3:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][3]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    elif index == 5:
                        new_file[f"F{z}"] = copy
                    elif index == 6:
                        new_file[f"G{z}"] = copy
                    elif index == 7:
                        new_file[f"H{z}"] = copy
                    elif index == 8:
                        new_file[f"I{z}"] = copy
                    elif index == 9:
                        new_file[f"J{z}"] = copy
                    elif index == 10:
                        new_file[f"K{z}"] = copy
                    z += 1
            if value == 4:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    elif index == 5:
                        new_file[f"F{z}"] = copy
                    elif index == 6:
                        new_file[f"G{z}"] = copy
                    elif index == 7:
                        new_file[f"H{z}"] = copy
                    elif index == 8:
                        new_file[f"I{z}"] = copy
                    elif index == 9:
                        new_file[f"J{z}"] = copy
                    elif index == 10:
                        new_file[f"K{z}"] = copy
                    z += 1
            if value == 5:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    elif index == 5:
                        new_file[f"F{z}"] = copy
                    elif index == 6:
                        new_file[f"G{z}"] = copy
                    elif index == 7:
                        new_file[f"H{z}"] = copy
                    elif index == 8:
                        new_file[f"I{z}"] = copy
                    elif index == 9:
                        new_file[f"J{z}"] = copy
                    elif index == 10:
                        new_file[f"K{z}"] = copy
                    z += 1

            if value == 6:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    elif index == 5:
                        new_file[f"F{z}"] = copy
                    elif index == 6:
                        new_file[f"G{z}"] = copy
                    elif index == 7:
                        new_file[f"H{z}"] = copy
                    elif index == 8:
                        new_file[f"I{z}"] = copy
                    elif index == 9:
                        new_file[f"J{z}"] = copy
                    elif index == 10:
                        new_file[f"K{z}"] = copy
                    z += 1
            if value == 7:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    elif index == 5:
                        new_file[f"F{z}"] = copy
                    elif index == 6:
                        new_file[f"G{z}"] = copy
                    elif index == 7:
                        new_file[f"H{z}"] = copy
                    elif index == 8:
                        new_file[f"I{z}"] = copy
                    elif index == 9:
                        new_file[f"J{z}"] = copy
                    elif index == 10:
                        new_file[f"K{z}"] = copy
                    z += 1
            if value == 8:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    elif index == 5:
                        new_file[f"F{z}"] = copy
                    elif index == 6:
                        new_file[f"G{z}"] = copy
                    elif index == 7:
                        new_file[f"H{z}"] = copy
                    elif index == 8:
                        new_file[f"I{z}"] = copy
                    elif index == 9:
                        new_file[f"J{z}"] = copy
                    elif index == 10:
                        new_file[f"K{z}"] = copy
                    z += 1
            if value == 9:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    elif index == 5:
                        new_file[f"F{z}"] = copy
                    elif index == 6:
                        new_file[f"G{z}"] = copy
                    elif index == 7:
                        new_file[f"H{z}"] = copy
                    elif index == 8:
                        new_file[f"I{z}"] = copy
                    elif index == 9:
                        new_file[f"J{z}"] = copy
                    elif index == 10:
                        new_file[f"K{z}"] = copy
                    z += 1

            if value == 10:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    elif index == 5:
                        new_file[f"F{z}"] = copy
                    elif index == 6:
                        new_file[f"G{z}"] = copy
                    elif index == 7:
                        new_file[f"H{z}"] = copy
                    elif index == 8:
                        new_file[f"I{z}"] = copy
                    elif index == 9:
                        new_file[f"J{z}"] = copy
                    elif index == 10:
                        new_file[f"K{z}"] = copy
                    z += 1

            if value ==11:
                for x in range(len(datas)):
                    # Copiar los datos de la fila especificada
                    copy = datas[x][value]
                    # Asignar los datos copiados a la nueva hoja de cálculo
                    if index == 0:
                        new_file[f"A{z}"] = copy
                    elif index == 1:
                        new_file[f"B{z}"] = copy
                    elif index == 2:
                        new_file[f"C{z}"] = copy
                    elif index == 3:
                        new_file[f"D{z}"] = copy
                    elif index == 4:
                        new_file[f"E{z}"] = copy
                    elif index == 5:
                        new_file[f"F{z}"] = copy
                    elif index == 6:
                        new_file[f"G{z}"] = copy
                    elif index == 7:
                        new_file[f"H{z}"] = copy
                    elif index == 8:
                        new_file[f"I{z}"] = copy
                    elif index == 9:
                        new_file[f"J{z}"] = copy
                    elif index == 10:
                        new_file[f"K{z}"] = copy
                    z += 1
        for i in range(len(filas_names)):
            if i == 0:
                new_file[f"A1"] = filas_names[i]
            if i == 1:
                new_file[f"B1"] = filas_names[i]
                i -= 1
            if i == 2:
                new_file[f"C1"] = filas_names[i]
                i -= 1
            if i == 3:
                new_file[f"D1"] = filas_names[i]
                i -= 1
            if i == 4:
                new_file[f"E1"] = filas_names[i]
                i -= 1
            if i == 5:
                new_file[f"F1"] = filas_names[i]
                i -= 1
            if i == 6:
                new_file[f"G1"] = filas_names[i]
                i -= 1
            if i == 7:
                new_file[f"H1"] = filas_names[i]
                i -= 1
            if i == 8:
                new_file[f"I1"] = filas_names[i]
                i -= 1
            if i == 9:
                new_file[f"J1"] = filas_names[i]
                i -= 1
            if i == 10:
                new_file[f"K1"] = filas_names[i]
                i -= 1
            if i == 11:
                new_file[f"L1"] = filas_names[i]
                i -= 1
            if i == 12:
                new_file[f"M1"] = filas_names[i]
                i -= 1
            if i == 13:
                new_file[f"N1"] = filas_names[i]
                i -= 1
            if i == 14:
                new_file[f"O1"] = filas_names[i]
                i -= 1
            if i == 15:
                new_file[f"P1"] = filas_names[i]
                i -= 1
            if i == 16:
                new_file[f"Q1"] = filas_names[i]
                i -= 1
            if i == 17:
                new_file[f"R1"] = filas_names[i]
                i -= 1
            if i == 18:
                new_file[f"S1"] = filas_names[i]
                i -= 1
            if i == 19:
                new_file[f"T1"] = filas_names[i]
                i -= 1
            if i == 20:
                new_file[f"U1"] = filas_names[i]
                i -= 1
            if i == 21:
                new_file[f"V1"] = filas_names[i]
                i -= 1
            if i == 22:
                new_file[f"W1"] = filas_names[i]
                i -= 1
            if i == 23:
                new_file[f"X1"] = filas_names[i]
                i -= 1
            if i == 24:
                new_file[f"Y1"] = filas_names[i]
                i -= 1
            if i == 25:
                new_file[f"Z1"] = filas_names[i]
                i -= 1
            if i == 26:
                new_file[f"AA1"] = filas_names[i]
                i -= 1

        # Guardar el archivo Excel de salida

        os.makedirs("archivos_creados", exist_ok=True)
        self.wb.save(f"archivos_creados/{filename}")

        # Cerrar el archivo Excel de salida
        self.wb.close()

        print("Archivo creado con éxito")


if __name__ == "__main__":

    train = False

    while train == False:

        # Crear la ventana principal

        # Botón para abrir el cuadro de diálogo de selección de archivo
        boton_seleccionar = tk.Button(text="Seleccionar Archivo xlsx", command=FileDialog())

        selected_file = FileDialog()

        archivo_xlsx = selected_file.archivo

        filename = input(
            "Nombre del archivo (solo el nombre, sin extensiones): ")

        title = input("Ponle un título a la hoja: ")

        row = (input(
            "Escribe el número de fila a copiar (si son múltiples, sepáralos con comas sin espacios): "))

        archivo = os.path.exists(filename + ".xlsx")

        if archivo == True:
            print("El archivo ya existe, por favor ingresa otro nombre")
        elif archivo_xlsx == False:
            print("No se seleccionó ningún archivo")
        else:
            train = True
    else:
        print("Creando archivo...")
        copy_obj = CopyFile(row=row, title=title, filename=filename, archivo=archivo_xlsx)
